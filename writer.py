#!/usr/bin/env python

from openpyxl import Workbook
from openpyxl import load_workbook


def write_to_cell(ws_active, cell_name, value):
	print cell_name,value
	ws_active[cell_name] = value


def save_to_file(wb, filename):
	wb.save(filename = filename)


"""
json_example:
{"A1":1,"A4":34}
"""
def json_to_excel(json,filename):
	try:
		wb = load_workbook(filename)
	except Exception,e:
		print e
		wb = Workbook()
	ws_active = wb.active
	for key in json:
		literal = ['A','B','C','D','E','F']
		index_key = literal.index(key[0])#key='A1',key[0]=A
		i=0
		for cell_value in json[key]:
			#for i in literal[index_key:len(json[key])]:
			literal_valid_index = literal.index(key[0])+i
			valid_key = literal[literal_valid_index]+key[1:]
				#print valid_key
			if cell_value:
				write_to_cell(ws_active,valid_key,cell_value)
			else:
				print "Cell empty"
			i+=1
	save_to_file(wb, filename)

"""
def main():
	json_example = {"A1":1,"A4":34}
	filename = "test.xlsx"
	json_to_excel(json_example, filename)

if __name__ == '__main__':
	main()
"""